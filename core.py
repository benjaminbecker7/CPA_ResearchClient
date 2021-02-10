#!/usr/local/bin/python

"""
527_Research_Utility.py
by Benjamin Becker
for the Center for Political Accountability

This python program allows for users to load previously recognized
category and public company IDs on their data spreadsheets for
categorical analysis of 527 data.
"""

import pandas as pd
from openpyxl import load_workbook

# Thought for optimization:
# This program should have a dedicated folder of csv files      <============== Done
# with all of the category labelings and the possible names of
# entities in those categories, and all of the public company
# IDs and the possible names of public companies under each
# ID.
# The program should load up those csv files at the start so it <============== Done
# has working dictionaries for entries.
# Users should be able to feed in filled-out Excel data to the  <============== Done
# program so the program can record new entries in the csv files
# to be able to more optimally fill in future forms.
# That way, when users want to update their own spreadsheets, 
# the program can take from the loaded dictionaries.

# Resource filepath constants
FILEPATH_CATEGORYID = "./resources/categoryIDs.txt"
FILEPATH_PUBLICID = "./resources/publicIDs.txt"

# Standardized layout constants
STD_SHEETNAME_DONATIONS = "DONATIONS"
STD_COLIDX_DONOR = "A"
STD_COLIDX_CATEGORYID = "H"
STD_COLIDX_PUBLICID = "I"


# TODO: REFACTOR TO HAVE DATASETS CONTAINING TUPLES OF DICT AND COLUMN LABEL
# Global dataset dictionaries
categoryID_data = {}
publicID_data = {}





# Loading Bar printer function
# (~~Credits to Greenstick (https://stackoverflow.com/questions/3173320/text-progress-bar-in-the-console)~~)
def printProgressBar (iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r"):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
        printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
    """
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    print("\r{prefix} |{bar}| {percent}% {suffix}".format(prefix=prefix, bar=bar, percent=percent, suffix=suffix), end = printEnd)
    # Print New Line on Complete
    if iteration == total: 
        print()
        
    return


#===============================================FILL SHEET MODULE===============================================

# Fill sheet at filepath with data from tuple of idsets
def fill_sheet(filepath, idsets, columns):
    try:
    
        # Open xlsx for read
        wb = load_workbook(filename=filepath)
        
        # Get the worksheet we need
        ws = wb[STD_SHEETNAME_DONATIONS]
        
        # Find number of idsets we're dealing with
        nsets = len(idsets)
        
        # Start at row 2
        active_row = 2
        
        goal = ws.max_row
        
        print()
        print("LOADING: Filling spreadsheet at filepath {}".format(filepath))
        
        
        while active_row <= goal:
            for set_idx in range(nsets):
                for key in idsets[set_idx].keys():
                    if ws[STD_COLIDX_DONOR + str(active_row)].value in idsets[set_idx][key]:
                        ws[str(columns[set_idx]) + str(active_row)] = key
            printProgressBar(active_row, goal)
            active_row += 1
        
        wb.save(filepath)
        
        input("DONE: Spreadsheet at filepath {} has been filled. Press ENTER to continue.".format(filepath))
        
    except Exception:
        input("WRITE ERROR: Could not write to spreadsheet at filepath {}. Press ENTER to continue.".format(filepath))
        return
        

def menu_fill_sheet():
    global categoryID_data
    global publicID_data
    while True:
        print()
        print("NOTE BEFORE UPDATING: Make sure the spreadsheet's columns are named as follows:")
        print("\tColumn A: Donor\t\t\t(Name of donor recorded for entry)")
        print("\tColumn H: Category\t\t(Category ID)")
        print("\tColumn I: Subcategory\t\t(Public ID or other identifier)")
        print("Also make sure that the donation data sheet is named DONATIONS.")
        print()
        print("What ID type would you like to fill?")
        print("[1] Category ID")
        print("[2] Public ID")
        print("[3] Both")
        print("[q] Go back")
        
        choice = input("Enter the number of your choice: ")
        if choice == '1':
            choice = input('Enter the filepath of the sheet you want to fill, or "." to go back: ')
            if choice == '.':
                continue
            fill_sheet(choice, (categoryID_data), ("Category"))
            
        elif choice == '2':
            choice = input('Enter the filepath of the sheet you want to fill, or "." to go back: ')
            if choice == '.':
                continue
            fill_sheet(choice, (publicID_data), ("Subcategory"))
        
        elif choice == '3':
            choice = input('Enter the filepath of the sheet you want to fill, or "." to go back: ')
            if choice == '.':
                continue
            fill_sheet(choice, (categoryID_data, publicID_data), (STD_COLIDX_CATEGORYID, STD_COLIDX_PUBLICID))
            
        elif choice == 'q':
            return
        else:
            print("Invalid input")



#===============================================================================================================



#===============================================ID INFO MODULE==================================================

# Search for an ID by entering in a name
def lookup_id(idset, name):
    for key in idset.keys():
        if name in idset[key]:
            print("{id_} is the ID for {name}".format(id_=key, name=name))
            input("Press ENTER to continue")
            return
    print('"{}" was not found in set'.format(name))
    input("Press ENTER to continue")
    return
    
# List names applicable to a given ID
def list_names(idset, id_):
    for key in idset.keys():
        if key == id_:
            print('These are the names with ID "{}"'.format(id_))
            for name in idset[key]:
                print("\t{}".format(name))
            input("Press ENTER to continue")
            return
    print('"{}" was not found in set'.format(id_))
    input("Press ENTER to continue")
    return
    
#List all IDs in given idset   
def list_ids(idset):
    for key in idset.keys():
        print("\t{}".format(key))
    input("Press ENTER to continue")
    return

# Choose what you would like to do
def menu_id_info_action(idset):
    while True:
        print()
        print("What would you like to view?")
        print("[1] Look up ID by name")
        print("[2] View all names for an ID")
        print("[3] List all IDs for type")
        print("[q] Go back")
        
        choice = input("Enter the number of your choice: ")
        if choice == '1':
            choice = input('Enter the name you would like to look up, or "." to go back: ')
            # check for back entry
            if choice == '.':
                continue
            lookup_id(idset, choice)
        elif choice == '2':
            choice = input('Enter the ID you would like to look up, or "." to go back: ')
            # check for back entry
            if choice == '.':
                continue
            list_names(idset, choice)
        elif choice == '3':
            list_ids(idset)
        elif choice == 'q':
            return
        else:
            print("Invalid input")

def menu_id_info():
    while True:
        print()
        print("What type of ID would you like to view?")
        print("[1] Category ID")
        print("[2] Public ID")
        print("[q] Go back")
        
        choice = input("Enter the number of your choice: ")
        if choice == '1':
            menu_id_info_action(categoryID_data)
        elif choice == '2':
            menu_id_info_action(publicID_data)
        elif choice == 'q':
            return
        else:
            print("Invalid input")

#===============================================================================================================





#=========================================RESOURCE WRITE MODULE=================================================

def write_to_resource(filepath, idset):
    print()
    
    goal = len(idset.keys())
    progress = 0;
    
    print("SAVING: Data to filepath {}".format(filepath))
    try:
        with open(filepath, 'w') as f:
            f.write("{}\n".format(goal))
            for key in idset.keys():

                # Format the write
                writestr = ""

                writestr += "{}~".format(key)

                for val in idset[key]:
                    writestr += "{}^".format(val)

                writestr = writestr.rstrip('^')
                writestr += '\n'

                # Write line
                f.write(writestr)

                progress += 1
                printProgressBar(progress, goal)
    except Exception:
        print("WRITE ERROR: Could not write to resource at filepath {}".format(filepath))
    return 

def write_data():
    global categoryID_data
    global publicID_data
    
    write_to_resource(FILEPATH_CATEGORYID, categoryID_data)
    write_to_resource(FILEPATH_PUBLICID, publicID_data)
    
    return

#===============================================================================================================





#=========================================UPDATE RESOURCES MODULE===============================================

"""
Takes data from 527 excel file at filepath and loads it into dictionaries
`
"""
def update_dictionaries_from_filepath(filepath):
    global categoryID_data
    global publicID_data
    
    try:
        dataframe = pd.read_excel(filepath)

        goal = len(dataframe.index) # Number of rows we need to read
        print()
        print("LOADING: Data from filepath {}".format(filepath))

        # Iterate through rows
        for progress, row in dataframe.iterrows():

            if pd.notna(row['Category']): 
                if not(row['Category'] in categoryID_data.keys()):
                    categoryID_data[row['Category']] = set()
                categoryID_data[row['Category']].add(row['Donor'])

            if (row['Category'] == "PUB") and pd.notna(row['Subcategory']):
                if not(row['Subcategory'] in publicID_data.keys()):
                    publicID_data[row['Subcategory']] = set()
                publicID_data[row['Subcategory']].add(row['Donor'])

            printProgressBar(progress + 1, goal)
    except Exception:
        print("LOAD ERROR: Could not load data from filepath {}".format(filepath))
    
    return

def menu_update_resources():
    print()
    print("NOTE BEFORE UPDATING: Make sure the spreadsheet's columns are named as follows:")
    print("\tColumn A: Donor\t\t\t(Name of donor recorded for entry)")
    print("\tColumn H: Category\t\t(Category ID)")
    print("\tColumn I: Subcategory\t\t(Public ID or other identifier)")
    print("Also make sure that the donation data sheet is the first sheet in the workbook.")
    filepath = input('Enter the path of the file you would like to submit, or "back" to go back: ')
    if filepath == "back":
        return
    
    update_dictionaries_from_filepath(filepath)
    
    print("Resources have been updated.")
    print("WARNING: New resources will not be saved until program has been properly exited")
    print("\tWould you like to save data now? (This could take some time)")
    choice = input("\t(y/n): ")
    if choice == 'y':
        write_data()
        
    print("Would you like to add another file?")
    choice = input("\t(y/n): ")
    if choice == 'y':
        menu_update_resources()
    return

#===============================================================================================================





#=========================================WELCOME SCREEN MODULE=================================================

# Displays the welcome menu
def menu_welcome():
    while(True):
        print()
        print("===============================================")
        print("Welcome to the 527 Research Utility!")
        print("What would you like to do?")
        print("[1] Fill ID data to a spreadsheet")
        print("[2] Get information about IDs")
        print("[3] Update resources (add ID data to dataset)")
        print("[q] Exit program")
        
        choice = input("Enter the number of your choice: ")
        if choice == '1':
            menu_fill_sheet()
        elif choice == '2':
            menu_id_info()
        elif choice == '3':
            menu_update_resources()
        elif choice == 'q':
            return
        else:
            print("Invalid input.")
            
#===============================================================================================================





#=========================================RESOURCE LOAD MODULE==================================================

"""
categoryIDs.txt and publicIDs.txt will be delimited as follows:
(1) Each line will describe an ID
(2) The first term in each line will be the ID and will be followed
    by a tilde ('~') to separate the ID from the list of possible names 
    it will denote.
(3) Following the tilde will be the possible names the ID will
    denote. Each possible name will be delimited by a carat ("^")
    because no company or individual will have a carat in their
    name
The only check on this that will be consistent will be for whether the tilde
has correctly delimited the file. This will be found by the function if
after the tilde split there are more than two items in the list.

FIRST LINE IN EACH FILE WILL BE NUMBER OF ENTRIES
"""
def read_resource_data(filepath, idset):
    print()
    
    # Load Category IDs
    try:
        with open(filepath, 'r') as f:
            progress = 0 # How many entries the function has loaded so far
            goal = 0 # How many entries the function will have to load
            for line_number, line in enumerate(f):
                # String the newline character
                line = line.rstrip('\n')

                # Get number of entries from first line
                if line_number == 0:
                    if line == '0':
                        print("No data in resource file {}".format(filepath))
                        f.close()
                        return
                    goal = int(line)

                    # Goal has been recorded, start load
                    print("LOADING: ID Data from path {}".format(filepath))
                    printProgressBar(progress, goal)
                    continue

                tilde_delimited = line.split('~') # Indicates line after it is split over tilde
                                                  # (I just really couldn't think of a better name)

                # TODO: MIGHT WANT TO CONSIDER NOT RETURNING BUT JUST SKIPPING THIS LOAD
                # Check if tilde split created more than 2 entries. If so, alert user and abort load.
                if len(tilde_delimited) > 2:
                    print()
                    print("LOAD ERROR: Multiple tildes on line {line_number} of file {file}".format(line_number=line_number, file=filepath))
                    print("Aborting load...")
                    f.close()
                    return

                # Check if key already in dictionary
                if tilde_delimited[0] in idset:
                    print()
                    print("LOAD ERROR: ID on line {line_number} of file {file} is a duplicate".format(line_number=line_number, file=filepath))
                    print("Aborting load...")
                    f.close()
                    return

                idset[tilde_delimited[0]] = set(tilde_delimited[1].split('^'))
                progress += 1
                printProgressBar(progress, goal)
    except Exception:
        print("LOAD ERROR: Could not load resource data from path {}".format(filepath))
    return

def load_data():
    global categoryID_data
    global publicID_data
    
    read_resource_data(FILEPATH_CATEGORYID, categoryID_data)
    read_resource_data(FILEPATH_PUBLICID, publicID_data)
    
    return

#===============================================================================================================
    
    
    

if __name__ == "__main__":
    load_data()
    menu_welcome()
    write_data()